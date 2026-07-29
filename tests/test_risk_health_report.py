"""风险健康报表（Phase 1 报表导出）回归测试。

覆盖：
- GET /v1/reports/risk-health/preview        预览 JSON（概览/项目风险/设备健康/Top）
- GET /v1/reports/risk-health/export?fmt=excel  下载 xlsx（PK 魔数 + 含「设备健康」sheet）
- GET /v1/reports/risk-health/export?fmt=pdf     下载 pdf（%PDF 魔数）
- 非法 period_type / fmt → 业务错误（HTTP 200 + body.code=400）

用真实库 + admin（数据范围=全部），前后清空 risk_health_snapshot 表，避免共享库残留干扰。
"""

import io
import secrets
from datetime import timedelta

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import delete

from app.core.database import SessionLocal
from app.main import app
from app.model.project import Project
from app.model.snapshot import RiskHealthSnapshot
from app.service.risk_health_report import period_bounds


@pytest.fixture
def client():
    with TestClient(app) as c:
        yield c


@pytest.fixture
def admin_token(client):
    r = client.post("/api/v1/auth/login", json={"username": "admin", "password": "Admin@123456"})
    assert r.status_code == 200, r.text
    return r.json()["data"]["access_token"]


@pytest.fixture
def wipe_snapshots():
    db = SessionLocal()
    try:
        db.execute(delete(RiskHealthSnapshot))
        db.commit()
    finally:
        db.close()
    yield
    db = SessionLocal()
    try:
        db.execute(delete(RiskHealthSnapshot))
        db.commit()
    finally:
        db.close()


def _seed(window: str, name_prefix: str):
    """在指定周期的当前/上一窗口各写一条项目快照 + 一条设备快照，返回 (pid, device_no)。"""
    db = SessionLocal()
    try:
        p = Project(name=f"{name_prefix}_报表", status="在建")
        db.add(p)
        db.flush()
        pid = p.id
        device_no = f"RHR_{name_prefix}_{secrets.token_hex(2)}"

        start, end, prev_start, prev_end = period_bounds(window)
        db.add(
            RiskHealthSnapshot(
                scope_type="project",
                ref_id=str(pid),
                name=f"{name_prefix}_报表",
                risk_index=72,
                risk_level="高",
                raw_score=150,
                snapshot_at=start + timedelta(hours=12),
            )
        )
        db.add(
            RiskHealthSnapshot(
                scope_type="project",
                ref_id=str(pid),
                name=f"{name_prefix}_报表",
                risk_index=60,
                risk_level="中",
                raw_score=120,
                snapshot_at=prev_start + timedelta(hours=12),
            )
        )
        db.add(
            RiskHealthSnapshot(
                scope_type="device",
                ref_id=device_no,
                name=f"{name_prefix}_设备",
                health_score=55,
                health_level="差",
                online_state="offline",
                snapshot_at=start + timedelta(hours=12),
            )
        )
        db.commit()
        return pid, device_no
    finally:
        db.close()


def test_preview_weekly(client, admin_token, wipe_snapshots):
    pid, device_no = _seed("weekly", "WEEK")
    h = {"Authorization": f"Bearer {admin_token}"}
    r = client.get(
        "/api/v1/reports/risk-health/preview", headers=h, params={"period_type": "weekly"}
    )
    assert r.status_code == 200, r.text
    d = r.json()["data"]
    assert d["period_type"] == "weekly"
    assert d["period_label"] == "周报（上周）"
    s = d["summary"]
    assert s["project_count"] >= 1
    assert s["device_count"] >= 1
    # 我建的项目应出现在 project_rows，且环比 Δ=+12（72-60）
    row = next((p for p in d["project_rows"] if p["name"] == "WEEK_报表"), None)
    assert row is not None, "预览应含自建项目"
    assert row["risk_index"] == 72
    assert row["prev_risk_index"] == 60
    assert row["delta"] == 12


def test_export_excel_has_device_sheet(client, admin_token, wipe_snapshots):
    _seed("weekly", "XLS")
    h = {"Authorization": f"Bearer {admin_token}"}
    r = client.get(
        "/api/v1/reports/risk-health/export",
        headers=h,
        params={"period_type": "weekly", "fmt": "excel"},
    )
    assert r.status_code == 200, r.text
    assert r.content[:2] == b"PK"
    wb = load_workbook(io.BytesIO(r.content))
    sheet_names = " ".join(wb.sheetnames)
    assert "设备健康" in sheet_names, f"Excel 应含设备健康明细 sheet，实际：{wb.sheetnames}"


def test_export_pdf(client, admin_token, wipe_snapshots):
    _seed("weekly", "PDF")
    h = {"Authorization": f"Bearer {admin_token}"}
    r = client.get(
        "/api/v1/reports/risk-health/export",
        headers=h,
        params={"period_type": "weekly", "fmt": "pdf"},
    )
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("application/pdf")
    assert r.content[:4] == b"%PDF"


def test_invalid_period_and_fmt(client, admin_token, wipe_snapshots):
    h = {"Authorization": f"Bearer {admin_token}"}
    # 非法周期
    r = client.get(
        "/api/v1/reports/risk-health/preview", headers=h, params={"period_type": "monthly"}
    )
    assert r.status_code == 200
    assert r.json()["code"] != 0, "非法 period_type 应返回业务错误"
    # 非法格式
    r = client.get(
        "/api/v1/reports/risk-health/export",
        headers=h,
        params={"period_type": "weekly", "fmt": "docx"},
    )
    assert r.status_code == 200
    assert r.json()["code"] != 0, "非法 fmt 应返回业务错误"
