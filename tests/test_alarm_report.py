"""告警报表 / 导出 回归测试。

覆盖：
- GET /v1/alarms/report        聚合统计 + 明细预览（时间范围 + 类型过滤 + 部门隔离）
- GET /v1/alarms/export?fmt=excel  下载 xlsx（zip 魔数 PK）
- GET /v1/alarms/export?fmt=pdf     下载 pdf（魔数 %PDF）
- 非法格式 → body.code=400（本项目 BusinessError 统一 HTTP 200 + code）

用真实库 + admin（数据范围=全部）；按项目清理自建数据。
"""

import io
import secrets
from datetime import datetime, timedelta, timezone

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import delete

from app.api.v1.alarms import _enumerate_periods, _parse_period
from app.core.database import SessionLocal
from app.main import app
from app.model.alarm import Alarm
from app.model.project import Project


def _uid() -> str:
    return secrets.token_hex(3)


@pytest.fixture
def client():
    with TestClient(app) as c:
        yield c


@pytest.fixture
def admin_token(client):
    r = client.post("/api/v1/auth/login", json={"username": "admin", "password": "Admin@123456"})
    assert r.status_code == 200, r.text
    return r.json()["data"]["access_token"]


def _headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture
def seeded():
    """建项目 + 3 条不同类型/处置状态的告警，返回 (project_id, uid)。"""
    u = _uid()
    db = SessionLocal()
    try:
        p = Project(name=f"RPT{u}_报表", status="在建")
        db.add(p)
        db.flush()
        pid = p.id
        now = datetime.now(timezone.utc)
        specs = [
            ("fence_intrusion", "严重", "待处理", now - timedelta(hours=1)),
            ("distance_too_close", "警告", "已处理", now - timedelta(hours=2)),
            ("device_alarm", "提示", "已消警", now - timedelta(days=1)),
        ]
        for i, (t, lvl, hs, at) in enumerate(specs):
            db.add(
                Alarm(
                    project_id=pid,
                    alarm_type=t,
                    device_type="locate",
                    device_name=f"设备{i}",
                    device_no=f"RPTDN{u}_{i}",
                    alarm_info=f"RPT{u} 告警{i}",
                    alarm_status="告警开始",
                    alarm_level=lvl,
                    handle_status=hs,
                    alarm_time=at,
                )
            )
        db.commit()
        yield pid, u
    finally:
        db.execute(delete(Alarm).where(Alarm.project_id == pid))
        db.execute(delete(Project).where(Project.id == pid))
        db.commit()
        db.close()


def test_alarm_report_summary(client, admin_token, seeded):
    pid, u = seeded
    r = client.get(
        "/api/v1/alarms/report",
        headers=_headers(admin_token),
        params={"project_id": pid},
    )
    assert r.status_code == 200, r.text
    data = r.json()["data"]
    summary = data["summary"]
    assert summary["total"] == 3
    assert summary["handled"] == 2  # 已处理 + 已消警
    assert summary["pending"] == 1
    # 按类型分布应含三类，且带中文 label
    labels = {it["label"] for it in summary["by_type"]}
    assert {"围栏侵入", "间距过近", "设备自报"} <= labels
    assert len(data["items"]) == 3
    # by_day 每项应含按类型/级别的日期细分（供堆叠柱状图）
    assert summary["by_day"], "by_day 不应为空"
    day0 = summary["by_day"][0]
    assert set(day0.keys()) >= {"date", "count", "by_type", "by_level"}
    assert sum(day0["by_type"].values()) == day0["count"]
    assert sum(day0["by_level"].values()) == day0["count"]


def test_alarm_daily(client, admin_token, seeded):
    """柱状图下钻：按 by_day 的每一天拉当日明细，数量与时间应一致。"""
    pid, u = seeded
    rep = client.get(
        "/api/v1/alarms/report", headers=_headers(admin_token), params={"project_id": pid}
    )
    days = rep.json()["data"]["summary"]["by_day"]
    assert days, "by_day 不应为空"
    for day in days:
        r = client.get(
            "/api/v1/alarms/daily",
            headers=_headers(admin_token),
            params={"project_id": pid, "date": day["date"]},
        )
        assert r.status_code == 200, r.text
        d = r.json()["data"]
        assert d["date"] == day["date"]
        assert d["total"] == day["count"]
        for it in d["items"]:
            assert (it["alarm_time"] or "")[:10] == day["date"]


def test_alarm_daily_invalid_date(client, admin_token, seeded):
    pid, u = seeded
    r = client.get(
        "/api/v1/alarms/daily",
        headers=_headers(admin_token),
        params={"project_id": pid, "date": "2026-13-40"},
    )
    # BusinessError → HTTP 200 + body.code=400（本项目校验约定）
    assert r.json()["code"] == 400, r.text


def test_alarm_report_summary_only(client, admin_token, seeded):
    """summary_only=true 只返回聚合统计，不构建明细预览（live 趋势用）。"""
    pid, u = seeded
    r = client.get(
        "/api/v1/alarms/report",
        headers=_headers(admin_token),
        params={"project_id": pid, "summary_only": True},
    )
    assert r.status_code == 200, r.text
    d = r.json()["data"]
    assert d["summary"]["total"] == 3
    assert d["items"] == []
    assert d["preview_count"] == 0
    assert d["filters_desc"]


def test_alarm_report_type_filter(client, admin_token, seeded):
    pid, u = seeded
    r = client.get(
        "/api/v1/alarms/report",
        headers=_headers(admin_token),
        params={"project_id": pid, "alarm_type": "fence_intrusion"},
    )
    assert r.status_code == 200, r.text
    summary = r.json()["data"]["summary"]
    assert summary["total"] == 1
    assert summary["by_type"][0]["key"] == "fence_intrusion"


def test_alarm_export_excel(client, admin_token, seeded):
    pid, u = seeded
    r = client.get(
        "/api/v1/alarms/export",
        headers=_headers(admin_token),
        params={"project_id": pid, "fmt": "excel"},
    )
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers["content-type"]
    assert "attachment" in r.headers["content-disposition"]
    # xlsx 是 zip：以 PK 魔数开头
    assert r.content[:2] == b"PK"
    assert len(r.content) > 1000


def test_alarm_export_pdf(client, admin_token, seeded):
    pid, u = seeded
    r = client.get(
        "/api/v1/alarms/export",
        headers=_headers(admin_token),
        params={"project_id": pid, "fmt": "pdf"},
    )
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("application/pdf")
    assert r.content[:4] == b"%PDF"
    assert len(r.content) > 1000


def test_alarm_export_invalid_format(client, admin_token, seeded):
    pid, u = seeded
    r = client.get(
        "/api/v1/alarms/export",
        headers=_headers(admin_token),
        params={"project_id": pid, "fmt": "csv"},
    )
    # BusinessError → HTTP 200 + body.code=400
    assert r.json()["code"] == 400, r.text


def test_alarm_export_daily_scope(client, admin_token, seeded):
    """导出按当日边界锁定，应与柱状图下钻 /daily 数量一致（下钻导出当日前后端契约）。"""
    pid, u = seeded
    rep = client.get(
        "/api/v1/alarms/report", headers=_headers(admin_token), params={"project_id": pid}
    )
    days = rep.json()["data"]["summary"]["by_day"]
    assert days, "by_day 不应为空"
    day = days[0]["date"]
    # 下钻当日明细数量
    d = client.get(
        "/api/v1/alarms/daily",
        headers=_headers(admin_token),
        params={"project_id": pid, "date": day},
    ).json()["data"]
    n = d["total"]
    # 用相同当日边界导出 excel，解析明细 sheet 行数应一致
    r = client.get(
        "/api/v1/alarms/export",
        headers=_headers(admin_token),
        params={
            "project_id": pid,
            "fmt": "excel",
            "start": f"{day}T00:00:00",
            "end": f"{day}T23:59:59.999999",
        },
    )
    assert r.status_code == 200, r.text
    assert r.content[:2] == b"PK"
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["明细"]
    data_rows = ws.max_row - 1  # 扣除表头
    assert data_rows == n, f"导出当日明细行数 {data_rows} 应与下钻数量 {n} 一致"


def test_alarm_export_period_scope(client, admin_token, seeded):
    """导出支持 granularity+period（一键整周/整月）：导出明细行数应 == /period 下钻数量。"""
    pid, u = seeded
    for gran in ("week", "month"):
        rep = client.get(
            "/api/v1/alarms/report",
            headers=_headers(admin_token),
            params={"project_id": pid, "granularity": gran},
        )
        periods = rep.json()["data"]["summary"]["by_period"]
        assert periods, f"{gran} by_period 不应为空"
        period = periods[0]["period"]
        # /period 下钻数量
        drill = client.get(
            "/api/v1/alarms/period",
            headers=_headers(admin_token),
            params={"project_id": pid, "granularity": gran, "period": period},
        ).json()["data"]
        n = drill["total"]
        # 用相同 granularity+period 导出 excel，明细行数应一致
        r = client.get(
            "/api/v1/alarms/export",
            headers=_headers(admin_token),
            params={
                "project_id": pid,
                "fmt": "excel",
                "granularity": gran,
                "period": period,
            },
        )
        assert r.status_code == 200, r.text
        assert r.content[:2] == b"PK"
        # 文件名应带 period 标识
        assert period in r.headers.get("content-disposition", ""), r.headers
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb["明细"]
        data_rows = ws.max_row - 1
        assert data_rows == n, f"{gran} 导出明细行数 {data_rows} 应与下钻数量 {n} 一致"


def test_alarm_export_period_invalid(client, admin_token, seeded):
    """非法周期值导出 → body.code=400。"""
    pid, u = seeded
    r = client.get(
        "/api/v1/alarms/export",
        headers=_headers(admin_token),
        params={
            "project_id": pid,
            "fmt": "excel",
            "granularity": "week",
            "period": "2026-W99",
        },
    )
    assert r.json()["code"] == 400, r.text


def test_alarm_report_granularity(client, admin_token, seeded):
    """granularity=week/month 时 by_period 按粒度聚合，且 by_day 始终按天（向后兼容）。"""
    pid, u = seeded
    rep = client.get(
        "/api/v1/alarms/report",
        headers=_headers(admin_token),
        params={"project_id": pid, "granularity": "week"},
    )
    assert rep.status_code == 200, rep.text
    summary = rep.json()["data"]["summary"]
    # by_day 始终按天分布，不随 granularity 改变
    assert len(summary["by_day"]) >= 1
    # by_period 按周聚合，周期总数应等于告警总数
    periods = summary["by_period"]
    assert periods, "by_period 不应为空"
    assert sum(p["count"] for p in periods) == 3
    for p in periods:
        assert set(p.keys()) >= {"period", "count", "by_type", "by_level"}
        assert sum(p["by_type"].values()) == p["count"]
        assert sum(p["by_level"].values()) == p["count"]

    # month 粒度
    rep_m = client.get(
        "/api/v1/alarms/report",
        headers=_headers(admin_token),
        params={"project_id": pid, "granularity": "month"},
    )
    pm = rep_m.json()["data"]["summary"]["by_period"]
    assert sum(p["count"] for p in pm) == 3


def test_alarm_report_granularity_invalid(client, admin_token, seeded):
    pid, u = seeded
    r = client.get(
        "/api/v1/alarms/report",
        headers=_headers(admin_token),
        params={"project_id": pid, "granularity": "year"},
    )
    assert r.json()["code"] == 400, r.text


def test_alarm_period_drilldown(client, admin_token, seeded):
    """周期下钻：week 模式每个 period 应覆盖其全部明细，数量对齐且周值一致。"""
    pid, u = seeded
    rep = client.get(
        "/api/v1/alarms/report",
        headers=_headers(admin_token),
        params={"project_id": pid, "granularity": "week"},
    )
    periods = rep.json()["data"]["summary"]["by_period"]
    for p in periods:
        r = client.get(
            "/api/v1/alarms/period",
            headers=_headers(admin_token),
            params={"project_id": pid, "granularity": "week", "period": p["period"]},
        )
        assert r.status_code == 200, r.text
        d = r.json()["data"]
        assert d["granularity"] == "week"
        assert d["period"] == p["period"]
        assert d["total"] == p["count"]
        for it in d["items"]:
            iso = datetime.fromisoformat(it["alarm_time"][:19]).isocalendar()
            assert f"{iso[0]}-W{iso[1]:02d}" == p["period"]


def test_alarm_period_month_and_day(client, admin_token, seeded):
    pid, u = seeded
    # month 下钻：回到当月全部 3 条
    rep = client.get(
        "/api/v1/alarms/report",
        headers=_headers(admin_token),
        params={"project_id": pid, "granularity": "month"},
    )
    month = rep.json()["data"]["summary"]["by_period"][0]["period"]
    r = client.get(
        "/api/v1/alarms/period",
        headers=_headers(admin_token),
        params={"project_id": pid, "granularity": "month", "period": month},
    )
    assert r.status_code == 200, r.text
    assert r.json()["data"]["total"] == 3
    # day 下钻：用 by_day 的首日
    day = rep.json()["data"]["summary"]["by_day"][0]["date"]
    rd = client.get(
        "/api/v1/alarms/period",
        headers=_headers(admin_token),
        params={"project_id": pid, "granularity": "day", "period": day},
    )
    assert rd.status_code == 200, rd.text
    for it in rd.json()["data"]["items"]:
        assert (it["alarm_time"] or "")[:10] == day


def test_alarm_period_invalid(client, admin_token, seeded):
    pid, u = seeded
    r = client.get(
        "/api/v1/alarms/period",
        headers=_headers(admin_token),
        params={"project_id": pid, "granularity": "week", "period": "2026-W99"},
    )
    assert r.json()["code"] == 400, r.text
    # 非法粒度
    r2 = client.get(
        "/api/v1/alarms/period",
        headers=_headers(admin_token),
        params={"project_id": pid, "granularity": "year", "period": "2026-07"},
    )
    assert r2.json()["code"] == 400, r2.text


def test_alarm_report_multipart_periods(client, admin_token):
    """多周期真实数据自证：跨多月的告警在 week/month 聚合下应呈现多个周期桶，
    且各桶 count 之和 == 总数（分桶口径正确、不漏不重）。"""
    db = SessionLocal()
    try:
        p = Project(name="MULTIPERIOD_临时", status="在建")
        db.add(p)
        db.flush()
        pid = p.id
        # 横跨 3 个月、不同日期/时刻，确保分入不同 ISO 周与不同月桶
        points = [
            (2026, 5, 10, 9),
            (2026, 5, 20, 14),
            (2026, 5, 28, 20),
            (2026, 6, 5, 10),
            (2026, 6, 15, 16),
            (2026, 6, 25, 8),
            (2026, 7, 3, 11),
            (2026, 7, 12, 13),
        ]
        for y, m, d, hh in points:
            db.add(
                Alarm(
                    project_id=pid,
                    alarm_type="fence_intrusion",
                    device_type="locate",
                    device_name="定位工牌",
                    device_no="MP-001",
                    alarm_info="HIST_SEED_MP",
                    alarm_status="告警开始",
                    alarm_level="严重",
                    handle_status="待处理",
                    alarm_time=datetime(y, m, d, hh, 0, 0, tzinfo=timezone.utc),
                )
            )
        db.commit()
        expected = len(points)

        # week 粒度：应跨多个 ISO 周
        rw = client.get(
            "/api/v1/alarms/report",
            headers=_headers(admin_token),
            params={"project_id": pid, "granularity": "week"},
        )
        assert rw.status_code == 200, rw.text
        weeks = rw.json()["data"]["summary"]["by_period"]
        week_keys = {p["period"] for p in weeks}
        assert len(week_keys) >= 3, f"week 桶数应>=3，实际 {week_keys}"
        assert sum(p["count"] for p in weeks) == expected

        # month 粒度：应恰好 3 个桶 2026-05 / 2026-06 / 2026-07
        rm = client.get(
            "/api/v1/alarms/report",
            headers=_headers(admin_token),
            params={"project_id": pid, "granularity": "month"},
        )
        assert rm.status_code == 200, rm.text
        months = rm.json()["data"]["summary"]["by_period"]
        month_keys = {p["period"] for p in months}
        assert month_keys == {"2026-05", "2026-06", "2026-07"}, month_keys
        assert sum(p["count"] for p in months) == expected
    finally:
        db.execute(delete(Alarm).where(Alarm.project_id == pid))
        db.execute(delete(Project).where(Project.id == pid))
        db.commit()
        db.close()


def test_alarm_export_snapshot_required_range(client, admin_token):
    """快照模式缺 start/end → body.code=400。"""
    r = client.get(
        "/api/v1/alarms/export",
        headers=_headers(admin_token),
        params={"fmt": "excel", "granularity": "month", "snapshot": True},
    )
    assert r.json()["code"] == 400, r.text


def test_alarm_export_snapshot_invalid_gran(client, admin_token):
    """快照粒度非法 → body.code=400。"""
    r = client.get(
        "/api/v1/alarms/export",
        headers=_headers(admin_token),
        params={
            "fmt": "excel",
            "granularity": "year",
            "snapshot": True,
            "start": "2026-05-01T00:00:00",
            "end": "2026-07-31T23:59:59",
        },
    )
    assert r.json()["code"] == 400, r.text


def test_alarm_export_snapshot_multi_sheet(client, admin_token):
    """跨周期历史快照：多 sheet 结构 + 每周期行数归属正确 + 合并行数==总数。

    复用一个跨 3 月、8 个时间点的数据集，验证 month/week 两种粒度快照。
    """
    db = SessionLocal()
    try:
        p = Project(name="SNAP_临时", status="在建")
        db.add(p)
        db.flush()
        pid = p.id
        points = [
            (2026, 5, 10, 9),
            (2026, 5, 20, 14),
            (2026, 5, 28, 20),
            (2026, 6, 5, 10),
            (2026, 6, 15, 16),
            (2026, 6, 25, 8),
            (2026, 7, 3, 11),
            (2026, 7, 12, 13),
        ]
        for y, m, d, hh in points:
            db.add(
                Alarm(
                    project_id=pid,
                    alarm_type="fence_intrusion",
                    device_type="locate",
                    device_name="定位工牌",
                    device_no="SNAP-001",
                    alarm_info="SNAP_SEED",
                    alarm_status="告警开始",
                    alarm_level="严重",
                    handle_status="待处理",
                    alarm_time=datetime(y, m, d, hh, 0, 0, tzinfo=timezone.utc),
                )
            )
        db.commit()
        expected = len(points)

        for gran in ("month", "week"):
            # 快照应覆盖 [start,end] 内的每一个周期（含空周期），与后端枚举口径一致
            expected_periods = set(
                _enumerate_periods(
                    gran,
                    datetime(2026, 5, 1, 0, 0, 0),
                    datetime(2026, 7, 31, 23, 59, 59),
                )
            )

            r = client.get(
                "/api/v1/alarms/export",
                headers=_headers(admin_token),
                params={
                    "project_id": pid,
                    "fmt": "excel",
                    "granularity": gran,
                    "snapshot": True,
                    "start": "2026-05-01T00:00:00",
                    "end": "2026-07-31T23:59:59",
                },
            )
            assert r.status_code == 200, r.text
            assert r.content[:2] == b"PK"
            wb = load_workbook(io.BytesIO(r.content))
            names = wb.sheetnames
            assert "概览" in names, names
            assert "明细合并" in names, names
            # 每周期一张明细 sheet（sheet 名==周期 key，本数据集无特殊字符）
            period_sheets = [
                n
                for n in names
                if n not in ("概览", "明细合并", "项目汇总") and not n.startswith("项目-")
            ]
            assert set(period_sheets) == expected_periods, (
                set(period_sheets),
                expected_periods,
            )
            # 按项目分 sheet：单一项目数据集应有「项目-{项目名}」明细 sheet，行数==总数（不漏不重）
            proj_sheets = [n for n in names if n.startswith("项目-")]
            assert proj_sheets, names
            proj_ws = wb[proj_sheets[0]]
            assert proj_ws.max_row - 1 == expected, proj_ws.max_row
            # 合并明细行数 == 总数（合并明细不含分组标题行）
            merged = wb["明细合并"]
            assert merged.max_row - 1 == expected, merged.max_row
            # 概览应含迷你趋势图（openpyxl 图表对象）
            overview = wb["概览"]
            assert len(overview._charts) >= 1, "概览应含迷你趋势图"
            # 每周期 sheet 的明细数据行（第2列有时间的行）全部落入该周期边界内
            per_sheet_total = 0
            for pk in period_sheets:
                b_start, b_end = _parse_period(gran, pk)
                ws = wb[pk]
                sheet_data = 0
                for ri in range(2, ws.max_row + 1):
                    t = ws.cell(row=ri, column=2).value  # 告警时间列
                    if not t:
                        # 分组标题行（▸ 项目名）第2列为空，跳过
                        continue
                    dt = datetime.fromisoformat(str(t)[:19])
                    assert b_start <= dt <= b_end, f"{pk}: {t} 落在边界外"
                    sheet_data += 1
                per_sheet_total += sheet_data
            # 各周期明细数据行之和 == 总数（不漏不重）
            assert per_sheet_total == expected, per_sheet_total
            # 项目汇总 sheet 存在且按项目计数合计 == 总数
            assert "项目汇总" in names, names
            proj_ws = wb["项目汇总"]
            total_proj = sum(
                (proj_ws.cell(row=ri, column=2).value or 0) for ri in range(2, proj_ws.max_row + 1)
            )
            assert total_proj == expected, total_proj

        # PDF 快照：魔数 %PDF
        rp = client.get(
            "/api/v1/alarms/export",
            headers=_headers(admin_token),
            params={
                "project_id": pid,
                "fmt": "pdf",
                "granularity": "month",
                "snapshot": True,
                "start": "2026-05-01T00:00:00",
                "end": "2026-07-31T23:59:59",
            },
        )
        assert rp.status_code == 200, rp.text
        assert rp.content[:4] == b"%PDF"
        assert len(rp.content) > 1000
    finally:
        db.execute(delete(Alarm).where(Alarm.project_id == pid))
        db.execute(delete(Project).where(Project.id == pid))
        db.commit()
        db.close()


def test_pdf_snapshot_trend_chart_embedded():
    """PDF 快照应内嵌「各周期告警数」趋势柱状图（reportlab 原生绘图）。

    验证：
    - `_make_period_bar_chart` 返回 Drawing 且数据/类目正确；
    - 含周期时生成的 PDF 明显大于空周期（图表矢量命令已写入）。
    """
    from app.service.alarm_report import (
        _make_period_bar_chart,
        build_pdf_snapshot,
    )

    keys = ["2026-05", "2026-06", "2026-07"]
    series = {
        "fence_intrusion": [40, 30, 120],
        "distance_too_close": [19, 26, 113],
        "device_alarm": [50, 50, 200],
    }
    d = _make_period_bar_chart(keys, series, "STSong-Light")
    # Drawing 是 platypus Flowable，可直接加入 PDF；内含 BarChart + Legend
    assert type(d).__name__ == "Drawing"
    chart = d.contents[0]
    # 堆叠：三类告警各为一条数据系列，且类目正确
    assert len(chart.data) == 3, chart.data
    assert list(chart.categoryAxis.categoryNames) == keys
    assert getattr(chart.categoryAxis, "style", None) == "stacked"
    # 每层颜色不同（分色）
    fills = [chart.bars[i].fillColor for i in range(3)]
    assert len({str(c) for c in fills}) == 3, fills
    # 颜色图例已绘制
    assert any(type(c).__name__ == "Legend" for c in d.contents)

    summary = {
        "total": 648,
        "handled": 206,
        "pending": 442,
        "handle_rate": 0.318,
    }
    period_rows = {
        pk: [
            {
                "project_id": 1,
                "alarm_type": "device_alarm",
                "alarm_level": "提示",
                "handle_status": "待处理",
                "alarm_time": "2026-06-10T10:00:00",
            }
        ]
        for pk in keys
    }
    meta = {"title": "t", "generated_at": "now", "filters_desc": "全部"}

    pdf_with_chart = build_pdf_snapshot("month", keys, period_rows, summary, meta, {1: "P"})
    pdf_empty = build_pdf_snapshot("month", [], {}, summary, meta, {1: "P"})
    assert pdf_with_chart[:4] == b"%PDF"
    assert pdf_empty[:4] == b"%PDF"
    # 多周期带趋势图，体积应显著大于无周期（仅文字表）
    assert len(pdf_with_chart) > len(pdf_empty) + 500, (
        len(pdf_with_chart),
        len(pdf_empty),
    )


def _fake_period_rows(per_type_map: dict) -> dict:
    """构造最小告警行，使 aggregate_alarms 能聚合出三类计数。"""
    rows: dict = {}
    for pk, per_type in per_type_map.items():
        pk_rows = []
        for t, n in per_type.items():
            for _ in range(n):
                pk_rows.append(
                    {
                        "alarm_type": t,
                        "alarm_level": "紧急",
                        "handle_status": "已处置",
                        "alarm_time": "2026-06-15T10:00:00",
                    }
                )
        rows[pk] = pk_rows
    return rows


def test_excel_snapshot_stacked_chart():
    """Excel 概览图应为「按类型分色堆叠」：stacked 分组、3 系列、红/橙/蓝、底部图例。"""
    from io import BytesIO

    from openpyxl import load_workbook

    from app.service.alarm_report import build_excel_snapshot

    keys = ["2026-05", "2026-06", "2026-07"]
    period_rows = _fake_period_rows(
        {
            "2026-05": {"fence_intrusion": 5, "distance_too_close": 3, "device_alarm": 2},
            "2026-06": {"fence_intrusion": 8, "distance_too_close": 1, "device_alarm": 4},
            "2026-07": {"fence_intrusion": 10, "distance_too_close": 5, "device_alarm": 7},
        }
    )
    summary = {"total": 45, "handled": 45, "pending": 0, "handle_rate": 1.0}
    meta = {"title": "t", "generated_at": "now", "filters_desc": "全部"}

    data = build_excel_snapshot("month", keys, period_rows, summary, meta)
    wb = load_workbook(BytesIO(data))
    ws = wb["概览"]
    assert len(ws._charts) == 1, ws._charts
    chart = ws._charts[0]
    # 堆叠分组（与 PDF/前端预览同源）
    assert chart.grouping == "stacked", chart.grouping
    assert type(chart).__name__ == "BarChart"
    # 三个数据系列：围栏侵入 / 间距过近 / 设备自报
    assert len(chart.series) == 3, len(chart.series)
    # 分类型上色（与 PDF/前端预览一致：红/橙/蓝）
    colors = [s.graphicalProperties.solidFill.srgbClr for s in chart.series]
    assert set(colors) == {"C00000", "ED7D31", "2E75B6"}, colors
    # 底部图例
    assert chart.legend is not None and chart.legend.position == "b"
