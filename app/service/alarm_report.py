"""告警报表导出：Excel（openpyxl）/ PDF（reportlab, 内置 CJK 字体）。

- build_excel(rows, summary, meta) -> bytes  两张表：概览 + 明细
- build_pdf(rows, summary, meta) -> bytes     标题 + 概览表 + 明细表（中文用 STSong-Light）

meta: {title, generated_at, filters_desc}
rows: query_alarms_for_report 的输出（已按数据范围过滤）
summary: aggregate_alarms 的输出
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.service.alarm_service import (
    aggregate_alarms,
    alarm_type_label,
)

# 明细列（key, 表头, 取值转换）
_COLUMNS: list[tuple[str, str]] = [
    ("id", "ID"),
    ("alarm_time", "告警时间"),
    ("alarm_type", "类型"),
    ("alarm_level", "级别"),
    ("device_type", "设备类型"),
    ("device_no", "设备编号"),
    ("device_name", "设备名称"),
    ("fence_name", "关联围栏"),
    ("alarm_info", "告警内容"),
    ("alarm_status", "告警状态"),
    ("handle_status", "处置状态"),
    ("work_plan_id", "作业计划ID"),
]

# 明细 sheet 表头样式（概览/快照复用）
_DETAIL_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_DETAIL_HEADER_FONT = Font(bold=True, color="FFFFFF")
_DETAIL_WIDTHS = {
    1: 8,
    2: 20,
    3: 12,
    4: 8,
    5: 12,
    6: 16,
    7: 16,
    8: 16,
    9: 40,
    10: 12,
    11: 12,
    12: 12,
}


def _cell(row: dict, key: str) -> Any:
    v = row.get(key)
    if key == "alarm_type":
        return alarm_type_label(v)
    if v is None:
        return ""
    return v


# Excel sheet 名规则：≤31 字符，禁用的字符 : \ / ? * [ ]
_INVALID_SHEET_CHARS = [":", "\\", "/", "?", "*", "[", "]"]


def _safe_sheet(name: str) -> str:
    s = name
    for ch in _INVALID_SHEET_CHARS:
        s = s.replace(ch, "-")
    s = s.strip()
    if not s:
        s = "周期"
    return s[:31]


def _write_detail_sheet(ws, rows: list[dict]) -> None:
    """写明细列（带表头样式 + 列宽 + 冻结首行）。供单表导出与快照每周期表复用。"""
    for ci, (_, label) in enumerate(_COLUMNS, start=1):
        c = ws.cell(row=1, column=ci, value=label)
        c.fill = _DETAIL_HEADER_FILL
        c.font = _DETAIL_HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(rows, start=2):
        for ci, (key, _) in enumerate(_COLUMNS, start=1):
            ws.cell(row=ri, column=ci, value=_cell(row, key))
    for ci, w in _DETAIL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"


def _write_detail_sheet_grouped(ws, rows: list[dict], project_names: dict | None = None) -> None:
    """每周期明细 sheet：按 project 分组，每组前加标题行（项目名 + 计数），数据行按明细列写出。"""
    for ci, (_, label) in enumerate(_COLUMNS, start=1):
        c = ws.cell(row=1, column=ci, value=label)
        c.fill = _DETAIL_HEADER_FILL
        c.font = _DETAIL_HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    ncols = len(_COLUMNS)

    def _pname(pid):
        if project_names and pid in project_names:
            return project_names[pid]
        return f"项目#{pid}" if pid is not None else "未关联项目"

    grouped: dict = {}
    for row in rows:
        grouped.setdefault(row.get("project_id"), []).append(row)
    keys = sorted(grouped.keys(), key=lambda k: (k is None, _pname(k)))

    ri = 2
    for pid in keys:
        g = grouped[pid]
        title = ws.cell(row=ri, column=1, value=f"▸ {_pname(pid)}（{len(g)} 条）")
        ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=ncols)
        title.font = Font(bold=True, color="FFFFFF")
        title.fill = PatternFill("solid", fgColor="2E75B6")
        title.alignment = Alignment(horizontal="left", vertical="center")
        ri += 1
        for row in g:
            for ci, (key, _) in enumerate(_COLUMNS, start=1):
                ws.cell(row=ri, column=ci, value=_cell(row, key))
            ri += 1

    for ci, w in _DETAIL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def build_excel(rows: list[dict], summary: dict, meta: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()

    # --- Sheet1: 概览 ---
    ws = wb.active
    ws.title = "概览"
    head_font = Font(bold=True, size=12)
    ws["A1"] = meta.get("title", "告警报表")
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"生成时间：{meta.get('generated_at', '')}"
    ws["A3"] = f"筛选条件：{meta.get('filters_desc', '全部')}"

    ws["A5"] = "汇总指标"
    ws["A5"].font = head_font
    kv = [
        ("告警总数", summary.get("total", 0)),
        ("已处置", summary.get("handled", 0)),
        ("待处理", summary.get("pending", 0)),
        ("处置率", f"{summary.get('handle_rate', 0) * 100:.1f}%"),
    ]
    r = 6
    for k, v in kv:
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
        r += 1

    def _dist_block(title: str, items: list[dict], key_field: str, r0: int) -> int:
        ws.cell(row=r0, column=1, value=title).font = head_font
        r0 += 1
        for it in items:
            label = it.get("label") or it.get(key_field) or it.get("date")
            ws.cell(row=r0, column=1, value=label)
            ws.cell(row=r0, column=2, value=it.get("count"))
            r0 += 1
        return r0 + 1

    r += 1
    r = _dist_block("按类型", summary.get("by_type", []), "key", r)
    r = _dist_block("按级别", summary.get("by_level", []), "key", r)
    r = _dist_block("按处置状态", summary.get("by_handle_status", []), "key", r)
    r = _dist_block("按日期", summary.get("by_day", []), "date", r)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16

    # --- Sheet2: 明细 ---
    ws2 = wb.create_sheet("明细")
    _write_detail_sheet(ws2, rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_excel_snapshot(
    granularity: str,
    period_keys: list[str],
    period_rows: dict[str, list[dict]],
    summary: dict,
    meta: dict,
    project_names: dict | None = None,
) -> bytes:
    """跨周期历史快照（多 sheet Excel）。

    结构：
    - 概览：汇总指标 + 按周期分布表（周期 | 总数 | 三类 | 待处理 | 已处置）+ 合计行
    - 每周期一张明细 sheet（sheet 名取周期 key，去重/截断合规）
    - 合并明细 sheet：首列「周期」，汇出全部周期告警，便于跨周期检索
    """
    from openpyxl import Workbook

    wb = Workbook()

    # --- Sheet1: 概览 ---
    ws = wb.active
    ws.title = "概览"
    head_font = Font(bold=True, size=12)
    ws["A1"] = meta.get("title", "告警历史快照")
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"生成时间：{meta.get('generated_at', '')}"
    ws["A3"] = f"筛选条件：{meta.get('filters_desc', '全部')}"
    ws["A4"] = f"快照粒度：{granularity}"
    ws["A5"] = (
        f"周期数：{len(period_keys)}"
        f"（{period_keys[0] if period_keys else '-'} ~ {period_keys[-1] if period_keys else '-'}）"
    )

    ws["A7"] = "汇总指标"
    ws["A7"].font = head_font
    kv = [
        ("告警总数", summary.get("total", 0)),
        ("已处置", summary.get("handled", 0)),
        ("待处理", summary.get("pending", 0)),
        ("处置率", f"{summary.get('handle_rate', 0) * 100:.1f}%"),
    ]
    r = 8
    for k, v in kv:
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
        r += 1

    # 按周期分布表
    r += 1
    ws.cell(row=r, column=1, value="按周期分布").font = head_font
    r += 1
    dist_header_row = r
    headers = ["周期", "总数", "围栏侵入", "间距过近", "设备自报", "待处理", "已处置"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=ci, value=h)
        c.fill = _DETAIL_HEADER_FILL
        c.font = _DETAIL_HEADER_FONT
    r += 1
    tot_acc = [0, 0, 0, 0, 0, 0]
    for pk in period_keys:
        rows_p = period_rows.get(pk, [])
        agg = aggregate_alarms(rows_p, granularity=granularity)
        counts = {it["key"]: it["count"] for it in agg["by_type"]}
        pending = agg["pending"]
        handled = agg["handled"]
        ws.cell(row=r, column=1, value=pk)
        ws.cell(row=r, column=2, value=len(rows_p))
        ws.cell(row=r, column=3, value=counts.get("fence_intrusion", 0))
        ws.cell(row=r, column=4, value=counts.get("distance_too_close", 0))
        ws.cell(row=r, column=5, value=counts.get("device_alarm", 0))
        ws.cell(row=r, column=6, value=pending)
        ws.cell(row=r, column=7, value=handled)
        tot_acc[0] += len(rows_p)
        tot_acc[1] += counts.get("fence_intrusion", 0)
        tot_acc[2] += counts.get("distance_too_close", 0)
        tot_acc[3] += counts.get("device_alarm", 0)
        tot_acc[4] += pending
        tot_acc[5] += handled
        r += 1
    dist_last_data_row = r - 1
    ws.cell(row=r, column=1, value="合计").font = Font(bold=True)
    for ci, v in enumerate(tot_acc, start=2):
        ws.cell(row=r, column=ci, value=v).font = Font(bold=True)
    ws.column_dimensions["A"].width = 22
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 12

    # 概览迷你趋势图：各周期告警数「按类型分色堆叠」柱状图（openpyxl 原生图表，与 PDF/前端预览同源）
    if period_keys:
        from openpyxl.chart import BarChart, Reference

        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.title = "各周期告警数 · 按类型堆叠"
        chart.height = 8
        chart.width = 16
        chart.style = 10
        # 数据系列：围栏侵入(列3)/间距过近(列4)/设备自报(列5)，表头行作系列名
        data = Reference(
            ws, min_col=3, max_col=5, min_row=dist_header_row, max_row=dist_last_data_row
        )
        cats = Reference(ws, min_col=1, min_row=dist_header_row + 1, max_row=dist_last_data_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        # 分类型上色：围栏侵入=红 / 间距过近=橙 / 设备自报=蓝（与 PDF/前端一致）
        _series_colors = ["C00000", "ED7D31", "2E75B6"]
        for _i, _s in enumerate(chart.series):
            _s.graphicalProperties.solidFill = _series_colors[_i % len(_series_colors)]
        chart.legend.position = "b"
        ws.add_chart(chart, "I7")

    # --- 每周期一张明细 sheet（按项目分组）---
    used: set[str] = set()
    for pk in period_keys:
        sname = _safe_sheet(pk)
        base = sname
        i = 2
        while sname in used:
            sname = f"{base[:27]}_{i}"
            i += 1
        used.add(sname)
        ws_p = wb.create_sheet(sname)
        _write_detail_sheet_grouped(ws_p, period_rows.get(pk, []), project_names)

    # --- 合并明细（首列周期）---
    ws_all = wb.create_sheet("明细合并")
    combined = [("__period__", "周期")] + _COLUMNS
    for ci, (_, label) in enumerate(combined, start=1):
        c = ws_all.cell(row=1, column=ci, value=label)
        c.fill = _DETAIL_HEADER_FILL
        c.font = _DETAIL_HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    ri = 2
    for pk in period_keys:
        for row in period_rows.get(pk, []):
            ws_all.cell(row=ri, column=1, value=pk)
            for ci, (key, _) in enumerate(_COLUMNS, start=2):
                ws_all.cell(row=ri, column=ci, value=_cell(row, key))
            ri += 1
    # 列宽：周期列 + 明细列
    ws_all.column_dimensions["A"].width = 16
    for ci, w in _DETAIL_WIDTHS.items():
        ws_all.column_dimensions[get_column_letter(ci + 1)].width = w
    ws_all.freeze_panes = "A2"

    # --- 项目汇总 sheet（跨整个窗口按项目聚合）---
    if project_names is not None:
        from collections import Counter

        proj_counter: Counter = Counter()
        proj_type: dict = {}
        proj_handle: dict = {}
        for r0 in (r for rows_ in period_rows.values() for r in rows_):
            pid = r0.get("project_id")
            name = project_names.get(pid, f"项目#{pid}" if pid is not None else "未关联项目")
            proj_counter[name] += 1
            proj_type.setdefault(name, Counter())[r0.get("alarm_type")] += 1
            proj_handle.setdefault(name, Counter())[r0.get("handle_status")] += 1
        total_all = sum(proj_counter.values())
        ws_proj = wb.create_sheet("项目汇总")
        pheaders = [
            "项目",
            "告警数",
            "占比",
            "围栏侵入",
            "间距过近",
            "设备自报",
            "待处理",
            "已处置",
        ]
        for ci, h in enumerate(pheaders, start=1):
            c = ws_proj.cell(row=1, column=ci, value=h)
            c.fill = _DETAIL_HEADER_FILL
            c.font = _DETAIL_HEADER_FONT
            c.alignment = Alignment(horizontal="center")
        rr = 2
        for name in sorted(proj_counter, key=lambda n: -proj_counter[n]):
            c = proj_counter[name]
            t = proj_type[name]
            h = proj_handle[name]
            pending = h.get("待处理", 0)
            ws_proj.cell(row=rr, column=1, value=name)
            ws_proj.cell(row=rr, column=2, value=c)
            ws_proj.cell(
                row=rr, column=3, value=f"{c / total_all * 100:.1f}%" if total_all else "0%"
            )
            ws_proj.cell(row=rr, column=4, value=t.get("fence_intrusion", 0))
            ws_proj.cell(row=rr, column=5, value=t.get("distance_too_close", 0))
            ws_proj.cell(row=rr, column=6, value=t.get("device_alarm", 0))
            ws_proj.cell(row=rr, column=7, value=pending)
            ws_proj.cell(row=rr, column=8, value=c - pending)
            rr += 1
        ws_proj.column_dimensions["A"].width = 28
        for col in "BCDEFGH":
            ws_proj.column_dimensions[col].width = 12
        ws_proj.freeze_panes = "A2"

    # --- 每项目一张明细 sheet（跨所有周期，按项目拆分）---
    if project_names is not None:
        per_proj: dict = {}
        for pk in period_keys:
            for row in period_rows.get(pk, []):
                name = _project_name(row.get("project_id"), project_names)
                per_proj.setdefault(name, []).append((pk, row))
        for name in sorted(per_proj, key=lambda n: -len(per_proj[n])):
            sname = _safe_sheet(f"项目-{name}")
            base = sname
            i = 2
            while sname in used:
                sname = f"{base[:25]}_{i}"
                i += 1
            used.add(sname)
            ws_pr = wb.create_sheet(sname)
            # 表头：周期 + 明细列
            proj_headers = [("__period__", "周期")] + _COLUMNS
            for ci, (_, label) in enumerate(proj_headers, start=1):
                c = ws_pr.cell(row=1, column=ci, value=label)
                c.fill = _DETAIL_HEADER_FILL
                c.font = _DETAIL_HEADER_FONT
                c.alignment = Alignment(horizontal="center")
            ri = 2
            for pk, row in sorted(per_proj[name], key=lambda x: x[0]):
                ws_pr.cell(row=ri, column=1, value=pk)
                for ci, (key, _) in enumerate(_COLUMNS, start=2):
                    ws_pr.cell(row=ri, column=ci, value=_cell(row, key))
                ri += 1
            ws_pr.column_dimensions["A"].width = 16
            for ci, w in _DETAIL_WIDTHS.items():
                ws_pr.column_dimensions[get_column_letter(ci + 1)].width = w
            ws_pr.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# 告警类型 → 颜色 / 中文标签（与前端概览表、预览图例保持一致）
_PERIOD_CHART_TYPE_ORDER = ["fence_intrusion", "distance_too_close", "device_alarm"]
_PERIOD_CHART_TYPE_COLORS = {
    "fence_intrusion": "#C00000",  # 围栏侵入 · 红
    "distance_too_close": "#ED7D31",  # 间距过近 · 橙
    "device_alarm": "#2E75B6",  # 设备自报 · 蓝
}
_PERIOD_CHART_TYPE_LABELS = {
    "fence_intrusion": "围栏侵入",
    "distance_too_close": "间距过近",
    "device_alarm": "设备自报",
}


def _make_period_bar_chart(
    period_keys: list[str],
    series: dict[str, list[int]],
    font: str,
    width: int = 560,
    height: int = 230,
) -> "object":
    """用 reportlab 原生绘图生成「各周期告警数（按类型分色堆叠）」柱状图。

    series: {类型键: 每个周期的计数列表}，按 _PERIOD_CHART_TYPE_ORDER 堆叠。
    周期较多（>12）时自动旋转类目标签，避免重叠；右上角绘制颜色图例。
    无额外依赖，与预览/Excel 同源。
    """
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.legends import Legend
    from reportlab.graphics.shapes import Drawing
    from reportlab.lib import colors as rl_colors

    n = len(period_keys)
    # 对齐到 period_keys 的二维数据（缺失类型补 0），顺序即堆叠由底到顶
    data = [list(series.get(k, [0] * n)) for k in _PERIOD_CHART_TYPE_ORDER]

    d = Drawing(width, height)
    chart = VerticalBarChart()
    chart.x = 42
    chart.y = 28
    chart.width = width - 72
    chart.height = height - 76
    chart.data = data
    chart.categoryAxis.categoryNames = list(period_keys)
    chart.categoryAxis.style = "stacked"
    chart.valueAxis.valueMin = 0
    flat = [v for s in data for v in s]
    vmax = max(flat) if flat else 0
    chart.valueAxis.valueMax = (vmax * 1.2) or 10
    chart.valueAxis.valueStep = max(1, int((vmax * 1.2) / 5)) or 1
    for i, k in enumerate(_PERIOD_CHART_TYPE_ORDER):
        chart.bars[i].fillColor = rl_colors.HexColor(_PERIOD_CHART_TYPE_COLORS[k])
        chart.bars[i].strokeColor = rl_colors.white
        chart.bars[i].strokeWidth = 0.3
    # 坐标轴标签使用 CJK 字体，避免数字/中文乱码
    chart.categoryAxis.labels.fontName = font
    chart.categoryAxis.labels.fontSize = 7 if n > 12 else 8
    chart.categoryAxis.labels.angle = 30 if n > 12 else 0
    chart.categoryAxis.labels.boxAnchor = "ne" if n > 12 else "n"
    chart.valueAxis.labels.fontName = font
    chart.valueAxis.labels.fontSize = 7
    d.add(chart)

    # 颜色图例（顶部水平排列，CJK 字体）
    legend = Legend()
    legend.x = 42
    legend.y = height - 8
    legend.dx = 8
    legend.dy = 8
    legend.fontName = font
    legend.fontSize = 7
    legend.alignment = "left"
    legend.boxAnchor = "nw"
    legend.columnMaximum = 1  # 每条类型独占一列 → 水平一行
    legend.deltax = 84
    legend.colorNamePairs = [
        (rl_colors.HexColor(_PERIOD_CHART_TYPE_COLORS[k]), _PERIOD_CHART_TYPE_LABELS[k])
        for k in _PERIOD_CHART_TYPE_ORDER
    ]
    d.add(legend)
    return d


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

_CJK_FONT = "STSong-Light"
_font_registered = False


def _ensure_cjk_font() -> str:
    """注册 reportlab 内置 CJK 字体（无需外部字体文件）。"""
    global _font_registered
    if not _font_registered:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont

        pdfmetrics.registerFont(UnicodeCIDFont(_CJK_FONT))
        _font_registered = True
    return _CJK_FONT


def build_pdf(rows: list[dict], summary: dict, meta: dict) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    font = _ensure_cjk_font()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=meta.get("title", "告警报表"),
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("cjk_title", parent=styles["Title"], fontName=font, fontSize=18)
    normal = ParagraphStyle("cjk_normal", parent=styles["Normal"], fontName=font, fontSize=9)
    small = ParagraphStyle("cjk_small", parent=styles["Normal"], fontName=font, fontSize=8)

    elems: list = [
        Paragraph(meta.get("title", "告警报表"), title_style),
        Spacer(1, 4 * mm),
        Paragraph(f"生成时间：{meta.get('generated_at', '')}", normal),
        Paragraph(f"筛选条件：{meta.get('filters_desc', '全部')}", normal),
        Spacer(1, 4 * mm),
    ]

    # 概览表
    ov = [
        ["告警总数", summary.get("total", 0), "已处置", summary.get("handled", 0)],
        [
            "待处理",
            summary.get("pending", 0),
            "处置率",
            f"{summary.get('handle_rate', 0) * 100:.1f}%",
        ],
    ]
    ov_tbl = Table(ov, colWidths=[35 * mm, 35 * mm, 35 * mm, 35 * mm])
    ov_tbl.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#DCE6F1")),
                ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#DCE6F1")),
            ]
        )
    )
    elems += [Paragraph("一、汇总指标", normal), Spacer(1, 2 * mm), ov_tbl, Spacer(1, 3 * mm)]

    # 按类型分布
    type_rows = [["告警类型", "数量"]] + [
        [it.get("label"), it.get("count")] for it in summary.get("by_type", [])
    ]
    type_tbl = Table(type_rows, colWidths=[60 * mm, 30 * mm])
    type_tbl.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ]
        )
    )
    elems += [Paragraph("二、按类型分布", normal), Spacer(1, 2 * mm), type_tbl, Spacer(1, 4 * mm)]

    # 明细表（最多渲染 500 行，避免超大 PDF）
    elems.append(Paragraph(f"三、告警明细（共 {len(rows)} 条，最多展示 500 条）", normal))
    elems.append(Spacer(1, 2 * mm))
    max_rows = 500
    header = [c[1] for c in _COLUMNS]
    data = [header]
    for row in rows[:max_rows]:
        data.append([Paragraph(str(_cell(row, k)), small) for k, _ in _COLUMNS])
    col_widths = [10, 26, 16, 12, 16, 22, 22, 22, 50, 16, 16, 16]
    col_widths = [w * mm for w in col_widths]
    detail_tbl = Table(data, colWidths=col_widths, repeatRows=1)
    detail_tbl.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FB")]),
            ]
        )
    )
    elems.append(detail_tbl)

    doc.build(elems)
    return buf.getvalue()


def build_pdf_snapshot(
    granularity: str,
    period_keys: list[str],
    period_rows: dict[str, list[dict]],
    summary: dict,
    meta: dict,
    project_names: dict | None = None,
) -> bytes:
    """跨周期历史快照（PDF 版）：概览 + 按周期分布表 + 合并明细（最多 500 行）。"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    font = _ensure_cjk_font()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=meta.get("title", "告警历史快照"),
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("cjk_title", parent=styles["Title"], fontName=font, fontSize=18)
    normal = ParagraphStyle("cjk_normal", parent=styles["Normal"], fontName=font, fontSize=9)
    small = ParagraphStyle("cjk_small", parent=styles["Normal"], fontName=font, fontSize=8)

    elems: list = [
        Paragraph(meta.get("title", "告警历史快照"), title_style),
        Spacer(1, 4 * mm),
        Paragraph(f"生成时间：{meta.get('generated_at', '')}", normal),
        Paragraph(f"筛选条件：{meta.get('filters_desc', '全部')}", normal),
        Paragraph(
            f"快照粒度：{granularity} · 周期数：{len(period_keys)}"
            f"（{period_keys[0] if period_keys else '-'} ~ {period_keys[-1] if period_keys else '-'}）",
            normal,
        ),
        Spacer(1, 4 * mm),
    ]

    # 概览表
    ov = [
        ["告警总数", summary.get("total", 0), "已处置", summary.get("handled", 0)],
        [
            "待处理",
            summary.get("pending", 0),
            "处置率",
            f"{summary.get('handle_rate', 0) * 100:.1f}%",
        ],
    ]
    ov_tbl = Table(ov, colWidths=[35 * mm, 35 * mm, 35 * mm, 35 * mm])
    ov_tbl.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#DCE6F1")),
                ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#DCE6F1")),
            ]
        )
    )
    elems += [Paragraph("一、汇总指标", normal), Spacer(1, 2 * mm), ov_tbl, Spacer(1, 3 * mm)]

    # 按周期分布表
    dist_header = ["周期", "总数", "围栏侵入", "间距过近", "设备自报", "待处理", "已处置"]
    dist_rows = [dist_header]
    for pk in period_keys:
        rows_p = period_rows.get(pk, [])
        agg = aggregate_alarms(rows_p, granularity=granularity)
        counts = {it["key"]: it["count"] for it in agg["by_type"]}
        dist_rows.append(
            [
                pk,
                len(rows_p),
                counts.get("fence_intrusion", 0),
                counts.get("distance_too_close", 0),
                counts.get("device_alarm", 0),
                agg["pending"],
                agg["handled"],
            ]
        )
    dist_tbl = Table(dist_rows, colWidths=[34 * mm] + [24 * mm] * 6, repeatRows=1)
    dist_tbl.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ]
        )
    )
    elems += [
        Paragraph("二、按周期分布", normal),
        Spacer(1, 2 * mm),
        dist_tbl,
        Spacer(1, 2 * mm),
    ]

    # 各周期告警数趋势图（按类型分色堆叠，reportlab 原生柱状图，与预览/Excel 同源）
    if period_keys:
        series = {
            "fence_intrusion": [row[2] for row in dist_rows[1:]],
            "distance_too_close": [row[3] for row in dist_rows[1:]],
            "device_alarm": [row[4] for row in dist_rows[1:]],
        }
        chart_drawing = _make_period_bar_chart(period_keys, series, font)
        elems += [
            Paragraph("（各周期告警数趋势 · 按类型分色堆叠）", small),
            chart_drawing,
            Spacer(1, 4 * mm),
        ]

    # 合并明细（最多 500 行）
    all_rows = [r for rows_ in period_rows.values() for r in rows_]
    elems.append(Paragraph(f"三、告警明细合并（共 {len(all_rows)} 条，最多展示 500 条）", normal))
    elems.append(Spacer(1, 2 * mm))
    max_rows = 500
    header = [c[1] for c in _COLUMNS]
    data = [header]
    for row in all_rows[:max_rows]:
        data.append([Paragraph(str(_cell(row, k)), small) for k, _ in _COLUMNS])
    col_widths = [10, 26, 16, 12, 16, 22, 22, 22, 50, 16, 16, 16]
    col_widths = [w * mm for w in col_widths]
    detail_tbl = Table(data, colWidths=col_widths, repeatRows=1)
    detail_tbl.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FB")]),
            ]
        )
    )
    elems.append(detail_tbl)

    # 按项目分布块（跨整个窗口按项目聚合）
    if project_names is not None:
        from collections import Counter

        proj_counter: Counter = Counter()
        proj_type: dict = {}
        for r0 in all_rows:
            pid = r0.get("project_id")
            name = project_names.get(pid, f"项目#{pid}" if pid is not None else "未关联项目")
            proj_counter[name] += 1
            proj_type.setdefault(name, Counter())[r0.get("alarm_type")] += 1
        total_all = len(all_rows)
        proj_header = ["项目", "告警数", "占比", "围栏侵入", "间距过近", "设备自报"]
        proj_rows = [proj_header]
        for name in sorted(proj_counter, key=lambda n: -proj_counter[n]):
            c = proj_counter[name]
            t = proj_type[name]
            proj_rows.append(
                [
                    name,
                    c,
                    f"{c / total_all * 100:.1f}%" if total_all else "0%",
                    t.get("fence_intrusion", 0),
                    t.get("distance_too_close", 0),
                    t.get("device_alarm", 0),
                ]
            )
        proj_tbl = Table(
            proj_rows,
            colWidths=[60 * mm, 24 * mm, 24 * mm, 24 * mm, 24 * mm, 24 * mm],
            repeatRows=1,
        )
        proj_tbl.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), font),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ]
            )
        )
        elems += [
            Spacer(1, 4 * mm),
            Paragraph("四、按项目分布", normal),
            Spacer(1, 2 * mm),
            proj_tbl,
        ]

    # 每项目明细（跨整个窗口，周期列前置；每项目最多 200 行）
    if project_names is not None:
        MAX_PP = 200
        per_proj: dict = {}
        for pk in period_keys:
            for r0 in period_rows.get(pk, []):
                name = _project_name(r0.get("project_id"), project_names)
                per_proj.setdefault(name, []).append((pk, r0))
        proj_detail_header = ["周期"] + [c[1] for c in _COLUMNS]
        proj_detail_widths = [16] + [10, 26, 16, 12, 16, 22, 22, 22, 50, 16, 16, 16]
        for name in sorted(per_proj, key=lambda n: -len(per_proj[n])):
            rows_p = per_proj[name]
            capped = len(rows_p) > MAX_PP
            note = (
                f"（共 {len(rows_p)} 条，最多展示 {MAX_PP} 条）"
                if capped
                else f"（共 {len(rows_p)} 条）"
            )
            elems.append(Spacer(1, 3 * mm))
            elems.append(Paragraph(f"五、按项目明细 · {name} {note}", normal))
            elems.append(Spacer(1, 1 * mm))
            data = [proj_detail_header]
            for pk, r0 in rows_p[:MAX_PP]:
                data.append([pk] + [Paragraph(str(_cell(r0, k)), small) for k, _ in _COLUMNS])
            tbl = Table(data, colWidths=[w * mm for w in proj_detail_widths], repeatRows=1)
            tbl.setStyle(
                TableStyle(
                    [
                        ("FONTNAME", (0, 0), (-1, -1), font),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.white, colors.HexColor("#F2F6FB")],
                        ),
                    ]
                )
            )
            elems.append(tbl)

    doc.build(elems)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 快照预览（JSON）：与 Excel/PDF 快照同源，供前端报表弹窗预览渲染
# ---------------------------------------------------------------------------


def _project_name(pid, project_names: dict | None) -> str:
    if project_names and pid in project_names:
        return project_names[pid]
    return f"项目#{pid}" if pid is not None else "未关联项目"


def build_snapshot_payload(
    granularity: str,
    period_keys: list[str],
    period_rows: dict[str, list[dict]],
    summary: dict,
    meta: dict,
    project_names: dict | None = None,
) -> dict:
    """把快照数据整理为 JSON 可序列化结构，供前端报表弹窗预览。

    与 build_excel_snapshot / build_pdf_snapshot 共用同一份 period_rows / summary /
    project_names，保证「预览即所见导出的内容」。
    """
    from collections import Counter

    periods = []
    for pk in period_keys:
        rows_p = period_rows.get(pk, [])
        agg = aggregate_alarms(rows_p, granularity=granularity)
        type_counts = {it["key"]: it["count"] for it in agg["by_type"]}
        level_counts = {it["key"]: it["count"] for it in agg["by_level"]}
        # 本周期按项目拆分（与明细 sheet 分组一致）
        proj_counter: Counter = Counter()
        proj_level: dict[str, Counter] = {}
        for r in rows_p:
            name = _project_name(r.get("project_id"), project_names)
            proj_counter[name] += 1
            proj_level.setdefault(name, Counter())[r.get("alarm_level")] += 1
        by_project = [
            {"project_name": n, "count": c, "by_level": dict(proj_level[n])}
            for n, c in sorted(proj_counter.items(), key=lambda x: -x[1])
        ]
        periods.append(
            {
                "period": pk,
                "total": len(rows_p),
                "by_type": type_counts,
                "by_level": level_counts,
                "pending": agg["pending"],
                "handled": agg["handled"],
                "by_project": by_project,
            }
        )

    # 项目汇总（跨整个窗口按项目聚合）—— 与「项目汇总」sheet 同源
    project_summary: list[dict] = []
    if project_names is not None:
        proj_counter = Counter()
        proj_type: dict[str, Counter] = {}
        proj_handle: dict[str, Counter] = {}
        for r0 in (r for rows_ in period_rows.values() for r in rows_):
            name = _project_name(r0.get("project_id"), project_names)
            proj_counter[name] += 1
            proj_type.setdefault(name, Counter())[r0.get("alarm_type")] += 1
            proj_handle.setdefault(name, Counter())[r0.get("handle_status")] += 1
        total_all = sum(proj_counter.values())
        for name in sorted(proj_counter, key=lambda n: -proj_counter[n]):
            c = proj_counter[name]
            t = proj_type[name]
            h = proj_handle[name]
            pending = h.get("待处理", 0)
            project_summary.append(
                {
                    "project_name": name,
                    "count": c,
                    "ratio": round(c / total_all, 4) if total_all else 0,
                    "by_type": {
                        "fence_intrusion": t.get("fence_intrusion", 0),
                        "distance_too_close": t.get("distance_too_close", 0),
                        "device_alarm": t.get("device_alarm", 0),
                    },
                    "pending": pending,
                    "handled": c - pending,
                }
            )

    # 每项目明细（与 Excel「按项目分 sheet」/ PDF「按项目明细」同源），供前端预览渲染
    projects_detail: list[dict] = []
    if project_names is not None:
        MAX_PP = 200
        per_proj: dict = {}
        for pk in period_keys:
            for r0 in period_rows.get(pk, []):
                name = _project_name(r0.get("project_id"), project_names)
                per_proj.setdefault(name, []).append((pk, r0))
        for name in sorted(per_proj, key=lambda n: -len(per_proj[n])):
            rows_p = per_proj[name]
            capped = len(rows_p) > MAX_PP
            detail = [
                {"period": pk, **{k: _cell(r0, k) for k, _ in _COLUMNS}}
                for pk, r0 in rows_p[:MAX_PP]
            ]
            projects_detail.append(
                {
                    "project_name": name,
                    "count": len(rows_p),
                    "capped": capped,
                    "rows": detail,
                }
            )

    return {
        "granularity": granularity,
        "period_keys": period_keys,
        "meta": meta,
        "summary": {
            "total": summary.get("total", 0),
            "handled": summary.get("handled", 0),
            "pending": summary.get("pending", 0),
            "handle_rate": summary.get("handle_rate", 0),
            "by_type": {it["key"]: it["count"] for it in summary.get("by_type", [])},
            "by_level": {it["key"]: it["count"] for it in summary.get("by_level", [])},
            "by_handle_status": {
                it["key"]: it["count"] for it in summary.get("by_handle_status", [])
            },
        },
        "periods": periods,
        "project_summary": project_summary,
        "projects_detail": projects_detail,
    }
