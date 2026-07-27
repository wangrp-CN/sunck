"""通用简表导出（报表导出对称化 ⑥）：任意业务列表 → Excel / PDF。

与 alarm_report.py 的重报表（多 sheet 快照/图表）互补，本模块面向
「隐患 / 设备」等常规列表提供轻量对称导出：

- build_simple_excel(columns, rows, meta, summary_blocks) -> bytes
    Sheet1 概览（标题/生成时间/筛选条件 + 分组统计块）+ Sheet2 明细。
- build_simple_pdf(columns, rows, meta, summary_blocks) -> bytes
    标题 + 概览统计 + 明细表（CJK 字体，最多 500 行）。

参数约定：
- columns: [(key, 表头, excel列宽, pdf列宽mm)]
- rows: list[dict]，值已是可展示字符串/数值（datetime 请上游先格式化）
- meta: {title, generated_at, filters_desc}
- summary_blocks: [(块标题, [(标签, 数量), ...])]
- image_columns: [(key, label, generator_fn)] —— generator_fn(row_dict) -> PNG bytes。
  单元格写为 ``<key>``；行单元格值调用 generator_fn(row) 取 PNG 嵌入。"""

from __future__ import annotations

import io
from collections.abc import Callable
from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.service.alarm_report import _ensure_cjk_font

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

Column = tuple[str, str, int, int]  # (key, label, excel_width, pdf_width_mm)
ImageColumn = tuple[str, str, Callable[[dict], bytes]]
"""(key, label, generator_fn)。generator_fn(row) -> PNG bytes。"""

_PDF_MAX_ROWS = 500


def _cell(row: dict, key: str) -> Any:
    v = row.get(key)
    return "" if v is None else v


def render_composite_sparkline_png(
    series_by_metric: dict[str, list[float]],
    colors: dict[str, str] | None = None,
    width: int = 240,
    height: int = 70,
    labels: dict[str, str] | None = None,
) -> bytes:
    """用 PIL 绘制项目级复合迷你趋势图（5 指标一条龙）。

    Args:
        series_by_metric: {metric_key: [v, v, ...]}（最多 5 条线）
        colors: {metric_key: '#RRGGBB'}；缺省用项目默认 5 色
        width/height: PNG 像素尺寸（建议 240x70，行高紧凑）
        labels: {metric_key: 中文短名}，左上角图例
    """
    from PIL import Image, ImageDraw

    palette = {
        "storm": "#409eff",
        "mttr": "#e6a23c",
        "dispatch_sla": "#67c23a",
        "hazard": "#13c2c2",
        "anomaly": "#9254de",
    }
    if colors:
        palette.update(colors)
    label_map = {
        "storm": "抑制",
        "mttr": "MTTR",
        "dispatch_sla": "SLA",
        "hazard": "闭环",
        "anomaly": "异常",
    }
    if labels:
        label_map.update(labels)

    img = Image.new("RGB", (width, height), (255, 255, 255))
    d = ImageDraw.Draw(img)
    pad_l, pad_r, pad_t, pad_b = 6, 6, 14, 4
    inner_w = max(1, width - pad_l - pad_r)
    inner_h = max(1, height - pad_t - pad_b)

    n = 1
    for vals in series_by_metric.values():
        n = max(n, len(vals))

    for key, vals in series_by_metric.items():
        if len(vals) < 1:
            continue
        color = palette.get(key, "#888")
        vmin = min(vals)
        vmax = max(vals)
        span = max(1e-6, vmax - vmin)
        m = len(vals)
        pts: list[tuple[float, float]] = []
        for i, v in enumerate(vals):
            x = pad_l + (inner_w * i / max(1, m - 1))
            y = pad_t + inner_h - ((v - vmin) / span) * inner_h
            pts.append((x, y))
        if len(pts) >= 2:
            d.line(pts, fill=color, width=1)
        if pts:
            ex, ey = pts[-1]
            d.ellipse((ex - 2, ey - 2, ex + 2, ey + 2), fill=color)

    # 左上图例
    lx = pad_l
    ly = 4
    for key in ("storm", "mttr", "dispatch_sla", "hazard", "anomaly"):
        if key not in series_by_metric:
            continue
        d.rectangle((lx, ly, lx + 8, ly + 8), fill=palette.get(key, "#888"))
        d.text((lx + 11, ly), label_map.get(key, key), fill="#303133")
        lx += 50

    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


def build_simple_excel(
    columns: list[Column],
    rows: list[dict],
    meta: dict,
    summary_blocks: list[tuple[str, list[tuple[str, Any]]]] | None = None,
    image_columns: list[ImageColumn] | None = None,
) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()

    # --- Sheet1: 概览 ---
    ws = wb.active
    ws.title = "概览"
    head_font = Font(bold=True, size=12)
    ws["A1"] = meta.get("title", "报表")
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"生成时间：{meta.get('generated_at', '')}"
    ws["A3"] = f"筛选条件：{meta.get('filters_desc', '全部')}"
    ws["A4"] = f"记录总数：{len(rows)}"

    r = 6
    for title, items in summary_blocks or []:
        ws.cell(row=r, column=1, value=title).font = head_font
        r += 1
        for label, count in items:
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=count)
            r += 1
        r += 1
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16

    # --- Sheet2: 明细 ---
    ws2 = wb.create_sheet("明细")
    img_keys = {key for key, _l, _fn in (image_columns or [])}
    all_cols: list[Column] = list(columns)
    for key, label, _fn in image_columns or []:
        all_cols.append((key, label, 34, 80))
    for ci, (_, label, _w, _pw) in enumerate(all_cols, start=1):
        c = ws2.cell(row=1, column=ci, value=label)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(rows, start=2):
        for ci, (key, _label, _w, _pw) in enumerate(all_cols, start=1):
            if key in img_keys:
                gen = next(fn for k, _l, fn in image_columns or [] if k == key)
                try:
                    png_bytes = gen(row)
                except Exception:
                    png_bytes = b""
                if png_bytes:
                    from openpyxl.drawing.image import Image as XLImage

                    xl_img = XLImage(io.BytesIO(png_bytes))
                    xl_img.width = 240
                    xl_img.height = 70
                    cell = ws2.cell(row=ri, column=ci)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    ws2.add_image(xl_img, cell.coordinate)
                    ws2.row_dimensions[ri].height = 56
            else:
                ws2.cell(row=ri, column=ci, value=_cell(row, key))
    for ci, (_key, _label, w, _pw) in enumerate(all_cols, start=1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_simple_pdf(
    columns: list[Column],
    rows: list[dict],
    meta: dict,
    summary_blocks: list[tuple[str, list[tuple[str, Any]]]] | None = None,
    image_columns: list[ImageColumn] | None = None,
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Image,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    font = _ensure_cjk_font()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=meta.get("title", "报表"),
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("cjk_title", parent=styles["Title"], fontName=font, fontSize=18)
    normal = ParagraphStyle("cjk_normal", parent=styles["Normal"], fontName=font, fontSize=9)
    small = ParagraphStyle("cjk_small", parent=styles["Normal"], fontName=font, fontSize=8)

    elems: list = [
        Paragraph(meta.get("title", "报表"), title_style),
        Spacer(1, 4 * mm),
        Paragraph(f"生成时间：{meta.get('generated_at', '')}", normal),
        Paragraph(f"筛选条件：{meta.get('filters_desc', '全部')}", normal),
        Paragraph(f"记录总数：{len(rows)}", normal),
        Spacer(1, 4 * mm),
    ]

    # 概览统计块（两列小表并排铺开）
    for title, items in summary_blocks or []:
        data = [[title, "数量"]] + [[label, count] for label, count in items]
        tbl = Table(data, colWidths=[60 * mm, 30 * mm])
        tbl.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), font),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ]
            )
        )
        elems += [tbl, Spacer(1, 3 * mm)]

    img_keys = {key for key, _l, _fn in (image_columns or [])}

    elems.append(Paragraph(f"明细（共 {len(rows)} 条，最多展示 {_PDF_MAX_ROWS} 条）", normal))
    elems.append(Spacer(1, 2 * mm))
    header = [label for _k, label, _w, _pw in columns]
    data = [header]
    for row in rows[:_PDF_MAX_ROWS]:
        row_cells = []
        for k, _label, _w, _pw in columns:
            if k in img_keys:
                gen = next(fn for kk, _l, fn in image_columns or [] if kk == k)
                try:
                    png_bytes = gen(row)
                except Exception:
                    png_bytes = b""
                if png_bytes:
                    img = Image(io.BytesIO(png_bytes), width=80 * mm, height=22 * mm)
                    row_cells.append(img)
                else:
                    row_cells.append(Paragraph("—", small))
            else:
                row_cells.append(Paragraph(str(_cell(row, k)), small))
        data.append(row_cells)
    col_widths = [pw * mm for _k, _label, _w, pw in columns]
    detail_tbl = Table(data, colWidths=col_widths, repeatRows=1)
    detail_tbl.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FB")]),
            ]
        )
    )
    elems.append(detail_tbl)

    doc.build(elems)
    return buf.getvalue()
